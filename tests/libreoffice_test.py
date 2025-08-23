#!/usr/bin/env python3
"""
File: libreoffice_test.py
Author: Bastian Cerf
Date: 19/05/2025
Description:
    Unit test of the LibreOffice module.

Company: Mecacerf SA
Website: http://mecacerf.ch
Contact: info@mecacerf.ch
"""

# Standard libraries
import pytest
from pathlib import Path
from typing import cast

# Third-party libraries
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell

# Internal libraries
from .test_constants import *
from core.spreadsheets.libreoffice import *


@pytest.fixture
def spreadsheet_file(arrange_assets: None) -> Path:
    """
    Creates a test spreadsheet file in the test assets folder.

    The spreadsheet contains a single sheet named "Test sheet" with numeric
    values in cells A1 to A3 and a formula in cell A4 that sums these values.
    The formula evaluates to 60, but no cached value is initially stored in
    the file.

    Returns:
        Path: Path to the created spreadsheet file.
    """
    test_file = Path(TEST_ASSETS_DST_FOLDER) / "test_wb.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws
    ws.title = "Test sheet"
    ws["A1"] = 10
    ws["A2"] = 20
    ws["A3"] = 30
    ws["A4"] = "=SUM(A1:A3)"
    wb.save(test_file)
    return test_file


def test_libreoffice_found():
    """
    The test currently relies on the LibreOffice module automatically
    finding the program on the filesystem, which may change in the future.
    """
    assert search_libreoffice() is not None


def test_evaluate_sheet(spreadsheet_file: Path):
    """
    Tests that a formula cell is correctly evaluated and its result stored.

    Opens the spreadsheet in `data_only` mode to verify that the initial formula
    result in cell A4 is `None` (i.e., not yet calculated). Then calls the
    evaluation function and reopens the file to check that the result has been
    evaluated and correctly updated to 60.
    """
    wb = load_workbook(spreadsheet_file, data_only=True)
    assert wb.active
    assert cast(Cell, wb.active["A4"]).value is None

    evaluate_calc(spreadsheet_file)

    wb = load_workbook(spreadsheet_file, data_only=True)
    assert wb.active
    assert cast(Cell, wb.active["A4"]).value == 60
