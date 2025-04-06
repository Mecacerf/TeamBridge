#!/usr/bin/env python3
"""
File: spreadsheet_time_tracker.py
Author: Bastian Cerf
Date: 21/02/2025
Description: 
    Implementation of the time tracker interface using spreadsheet files for data storage.

Company: Mecacerf SA
Website: http://mecacerf.ch
Contact: info@mecacerf.ch
"""

################################################################################################
#                           Note about spreadsheets management                                 #
# 
# - Library choice
# After some tests of the different available libraries for spreadsheet access, openpyxl appeared
# to be the most advanced one, with the only disadvantage that it supports only the XLSX format.
# 
# - Cells evaluation
# The software is designed to be optional, meaning the spreadsheets alone can be used to save and
# calculate employees information such as clock in / out events, monthly and daily balances, and 
# so on. That means that when an event is added in the spreadsheet by the software, the sheet
# must be reevaluated in order to update the affected cells values. The evaluation cannot be done
# directly by Python, since no formula interpreter seems to exist at this time. So the solution
# is to invoke libreoffice calc in headless mode (subprocess) and use it to save an evaluated 
# version of the sheet, then replace the old version by the new evaluated file.  
#
# - Spreadsheet file states
# According to the process described above, a spreadsheet can be in the following states:
#
# +---------------------+-------------+---------------------+--------------+-------------+
# |        State        | File System | Cached Cells Values | Write Access | Read Access |
# +---------------------+-------------+---------------------+--------------+-------------+
# | Saved and evaluated | Up to date  | Up to date          | Yes          | Yes         |
# | Modified in RAM     | Outdated    | Outdated            | Yes          | No          |
# | Saved on disk       | Up to date  | Unavailable         | Yes          | No          |
# | Evaluating          | Up to date  | Unavailable         | Yes          | No          |
# +---------------------+-------------+---------------------+--------------+-------------+
# 
# In the initial state, the file is saved on disk and the forumla cells have been evaluated, 
# meaning their values are available in the cells cache (when opnening the file in 
# data_only=True). 
# Once the employee's data have been modified, the file on disk is outdated as well as the 
# cells values (since the modified cell may be used in a formula). The read access is unset
# because the read values would be outdated.
# Once the file has been saved, changes are saved on disk but the file cache is deleted
# (this process is done automatically by openpyxl because it cannot evaluate the formulas). The
# read access is obviously still unavailable. 
# To evaluate the formula cells, libreoffice is invoked in a subprocess. The evaluation can take
# some time to finish. Once the evaluation is done, the old sheet file is replaced by the new
# one. The sheet is saved and evaluated again and the read access is available. Note that if
# a write is done during the sheet evaluation, it would be outdated on disk again.
# 
# When a file is initially opened (the Time Tracker is created), it is in the third state by
# default (saved on disk, data unavailable). This is because the formulas use the current
# date and time to calculate their values, and an evaluation is required to update these
# fields.
#
# - Design
# The problem arises when thinking about a common interface for both time trackers based on 
# spreadsheets or a database. With a database, read and write accesses would be faster and the 
# formula evaluation mechanism would probably be done on the fly with SQL commands.
# Nevertheless, two mechanisms can be identified:
# * Commit: after changes have been done, committing them would save the data in the file 
#   system. With the spreadsheets version that involves saving the file on disk. With a DB it 
#   would probably involve sending the queued SQL commands. This mechanism is implementation
#   dependant though. This function changes the state from "modified in RAM" to "saved on disk".
# * Evaluate: once changes have been committed, an evaluation must be done in order to calculate
#   the formula cells and read up-to-date data. A function can be used to know whether the read
#   functions are accessible or not at this moment. This function changes the state from "saved
#   on disk" to "evaluating" and finally "saved and evaluated".
#
# The state change functions can take a while to execute, depending on IO bound operations such
# as accessing the file system, a NAS or an external executable (libreoffice). It has been 
# decided that all functions will be blocking and synchronous and that a time tracker object
# will typically be used by a background thread.
################################################################################################

# Import openpyxl for spreadsheet manipulation
import openpyxl
import openpyxl.cell
import openpyxl.utils
import openpyxl.worksheet
# Import subprocess, pathlib and shutil that will be required to execute processes and manipulate files 
import subprocess
import pathlib
import shutil
import os
# Import the time tracker interface
from time_tracker_interface import ITodayTimeTracker, ClockEvent, ClockAction, IllegalReadException
# Import the spreadsheets database access
from spreadsheets_repository import SpreadsheetsRepository
# Import datetime
import datetime as dt
# Import logging module
import logging

################################################
#           Spreadsheet constants              #
################################################

## Sheets
# Init sheet index
SHEET_INIT = 0
# January sheet index, next months are incremented
SHEET_JANUARY = 1

## Init sheet
# Cell containing the name information
CELL_NAME = 'A6'
# Cell containing the firstname information
CELL_FIRSTNAME = 'A7'
# Cell containing the current date information
CELL_DATE = 'A8'
# Cell containing the current hour information
CELL_HOUR = 'A9'

## Month sheets
# Cell containing the monthly balance (total of daily balances)
CELL_MONTHLY_BALANCE = 'D8'
# Cell containing the first day of the month
CELL_FIRST_MONTH_DATE = 'B9'
# Cell containing the daily schedule for the first day of the month
CELL_DAILY_SCHEDULE = 'C9'
# Cell containing the daily balance for the first day of the month
CELL_DAILY_BALANCE = 'D9'
# Cell containing the worked time for the first day of the month
CELL_WORKED_TIME = 'E9'
# Column containing the first clock in action of the day
COL_FIRST_CLOCK_IN = 'G'
# Column containing the last clock out action of the day
COL_LAST_CLOCK_OUT = 'R'

################################################
#              General constants               #
################################################

# Path to 'soffice' in the filesystem
LIBREOFFICE_PATH = "C:\\Program Files\\LibreOffice\\program\\soffice"
# Temporary folder in which evaluated spreadsheets are placed by libreoffice
LIBREOFFICE_CACHE_FOLDER = ".tmp_calc/"

# Get file logger
LOGGER = logging.getLogger(__name__)

class SpreadsheetTimeTracker(ITodayTimeTracker):
    """
    Implementation of the time tracker that uses spreadsheet files as database.
    """

    def __init__(self, repository: SpreadsheetsRepository, employee_id: str, date: dt.date):
        """
        Open the employee's data for given date.

        Parameters:
            repository: spreadsheets repository access provider 
            id: employee unique ID
            date: date index
        Raise:
            ValueError: employee not found
            ValueError: wrong / unavailable date time                       TODO
            RuntimeError: employee's file is already opened
            OSError: error related to file access
            TimeoutError: failed to access database folder after timeout
            FileNotFoundError: employee's file not found
        """
        # Save current date
        self._date = date
        # Save employees database access provider
        self._repository = repository
        # Save employee's id
        self._employee_id = employee_id
        # Set readable flag to be initially False
        self._readable = False
        # Acquire employee's file
        self._file_path = self._repository.acquire_employee_file(self._employee_id)
        # Throw a file not found error if necessary
        if self._file_path is None:
            raise FileNotFoundError(f"File not found for employee's ID '{self._employee_id}'.")
        # Load workbook
        self.__load_workbook()
        # Even though the evaluated workbook might be available, it is defined that at time tracker
        # creation the data may be too old to be actually used. The time tracker must be evaluated
        # to access reading functions.
        self.__invalidate_eval_wb()

    def get_firstname(self) -> str:
        """
        Get employee's firstname.
        Always accessible.

        Returns:
            str: employee's firstname
        """
        # Get information in raw notebook
        return str(self._workbook_raw.worksheets[SHEET_INIT][CELL_FIRSTNAME].value)

    def get_name(self) -> str:
        """
        Get employee's name.
        Always accessible.

        Returns:
            str: employee's name
        """
        # Get information in raw notebook
        return str(self._workbook_raw.worksheets[SHEET_INIT][CELL_NAME].value)

    def __get_date_row(self) -> int:
        """
        Returns:
            int: the row containing the information related to current date
        """
        # Get coordinates of the first day of the month
        row, _ = openpyxl.utils.coordinate_to_tuple(CELL_FIRST_MONTH_DATE)
        # Find the row corresponding to current date
        return (row + self._date.day - 1)

    def __get_clock_action_for_cell(self, cell: openpyxl.cell.Cell) -> ClockAction:
        """
        Check if the given cell contains a clock in or a clock out action, based on its column index.
        """
        # Get first clock in column
        _, clock_in_col = openpyxl.utils.coordinate_to_tuple(f"{COL_FIRST_CLOCK_IN}1")
        # Get difference between first clock in column and given cell column
        delta_actions = cell.column - clock_in_col
        # Sanity check
        if delta_actions < 0:
            raise ValueError(f"Cannot evaluate clock action for cell '{cell.coordinate}' that is beyond limits.")
        # Since a clock in is followed by a clock out and so on, 0=in, 1=out, 2=in, etc
        if (delta_actions % 2) == 0:
            # This is a clock in column
            return ClockAction.CLOCK_IN
        else:
            # This is a clock out action
            return ClockAction.CLOCK_OUT

    def get_clock_events_today(self) -> list[ClockEvent]:
        """
        Get all clock-in/out events for the date.
        Always accessible.
        
        Returns:
            list[ClockEvent]: list of today's clock events (can be empty)
        """
        # Get current date row
        date_row = self.__get_date_row()
        # Get current month's sheet 
        month_sheet = self._workbook_raw.worksheets[self._date.month - 1 + SHEET_JANUARY]
        # Create current date events list
        clock_events = []
        # Get clock actions boundaries
        min_col, _, max_col, _ = openpyxl.utils.range_boundaries(f"{COL_FIRST_CLOCK_IN}1:{COL_LAST_CLOCK_OUT}1")
        # Iterate in clock actions row for current date
        for column in month_sheet.iter_cols(min_row=date_row, max_row=date_row, min_col=min_col, max_col=max_col):
            # Get cell value
            cell = column[0]
            # Check if the cell exists and contains an entry
            if cell and cell.value:
                # Create and append the clock event to the list
                event = ClockEvent(time=cell.value, action=self.__get_clock_action_for_cell(cell))
                clock_events.append(event)
            else:
                # No clock event in the cell
                # For this implementation, consider that no other clock action can exist after a hole is found
                break

        # Return the events list
        return clock_events
    
    def is_readable(self) -> bool:
        """
        Check if the reading functions are accessible at this moment. 
        They are initially not accessible (since the opened data are not evaluated) and
        get accessible after an evaluation is performed.

        Returns:
            bool: reading flag
        """
        return self._readable
    
    def __get_daily_info(self, first_cell: str) -> dt.timedelta:
        """
        Extract a daily information based on the given cell (column) and
        the current date (row). 

        Parameters:
            first_cell: a cell containing the information to identify the column
        Returns:
            timedelta: timedelta read in the identified cell
        Raises:
            IllegalReadException: read is unavailable  
        """
        # Check read status
        if not self.is_readable():
            raise IllegalReadException()
        # Get current date row
        row = self.__get_date_row()
        # Get column containing information
        _, column = openpyxl.utils.coordinate_to_tuple(first_cell)
        # Get current month's sheet in read mode
        month_sheet = self._workbook_eval.worksheets[self._date.month - 1 + SHEET_JANUARY]
        # Get the timedelta object
        timedelta = month_sheet.cell(row=row, column=column).value
        # It might happen that the object is a dt.time
        if isinstance(timedelta, dt.time):
            timedelta = dt.timedelta(hours=timedelta.hour, minutes=timedelta.minute, seconds=timedelta.second)
        return timedelta

    def get_daily_schedule(self) -> dt.timedelta:
        """
        Get employee's daily schedule (how much time he's supposed to work).
        Accessible when is_readable() returns True.

        Returns:
            timedelta: daily schedule
        """
        return self.__get_daily_info(CELL_DAILY_SCHEDULE)
    
    def get_daily_balance(self) -> dt.timedelta:
        """
        Get employee's daily balance (remaining time he's supposed to work).
        If the employee is clocked in the value is calculated based on the time the last evaluation
        has been done.
        Accessible when is_readable() returns True.

        Returns:
            timedelta: daily balance
        """
        return self.__get_daily_info(CELL_DAILY_BALANCE)

    def get_daily_worked_time(self) -> dt.timedelta:
        """
        Get employee's worked time for the day.
        If the employee is clocked in the value is calculated based on the time the last evaluation
        has been done.
        Accessible when is_readable() returns True.
        
        Returns:
            timedelta: delta time object
        """
        return self.__get_daily_info(CELL_WORKED_TIME)
    
    def get_monthly_balance(self) -> dt.timedelta:
        """
        Get employee's balance for the current month.
        Accessible when is_readable() returns True.

        Returns:
            timedelta: delta time object
        """
        # Check read status
        if not self.is_readable():
            raise IllegalReadException()
        # Get evaluated month's sheet 
        month_sheet = self._workbook_eval.worksheets[self._date.month - 1 + SHEET_JANUARY]
        # Get monthly balance value as a timedelta
        timedelta = month_sheet[CELL_MONTHLY_BALANCE].value
        # It might happen that the object is a dt.time
        if isinstance(timedelta, dt.time):
            timedelta = dt.timedelta(hours=timedelta.hour, minutes=timedelta.minute, seconds=timedelta.second)
        return timedelta

    def register_clock(self, event: ClockEvent) -> None:
        """
        Register a clock in/out event.
        After a clock event is registered, the reading functions are not available until a
        new evaluation is performed.

        Parameters:
            event: clock event object
        Raise:
            ValueError: double clock in/out detected
        """
        # Get current date cell
        date_row = self.__get_date_row()
        # Get current month's sheet
        # This function will only use the raw workbook
        month_sheet = self._workbook_raw.worksheets[self._date.month - 1 + SHEET_JANUARY]
        # Set written flag
        written = False
        # Get clock actions boundaries
        min_col, _, max_col, _ = openpyxl.utils.range_boundaries(f"{COL_FIRST_CLOCK_IN}1:{COL_LAST_CLOCK_OUT}1")
        # Previous clock action time, used to verify that clock event times are ascending
        prev_clock_time = None
        # Iterate in clock actions row for current date
        for column in month_sheet.iter_cols(min_row=date_row, max_row=date_row, min_col=min_col, max_col=max_col):
            # Get cell in the column
            cell = column[0]
            # Iterate until a free cell is found
            if cell.value:
                # Save previous clock action time
                prev_clock_time = cell.value
            else:
                # Cell value is None, free to write a clock action
                # For this implementation, check that the cell is intended to contain the given clock action and raise an error if not
                expected_action = self.__get_clock_action_for_cell(cell)
                if event.action != expected_action:
                    raise ValueError(f"Got a '{event.action}' while expecting a '{expected_action}'.")
                # Check that the clock action times are in ascending order
                if prev_clock_time and prev_clock_time > event.time:
                    raise ValueError(f"Cannot clock at {event.time} while last clock was at {prev_clock_time}.")
                # Write the clock action
                # Save as HH:MM, do not take seconds into account for final user simplicity
                cell.number_format = 'HH:MM'
                cell.value = dt.time(hour=event.time.hour, minute=event.time.minute, second=0)
                # Log action, set written flag and leave
                written = True
                LOGGER.debug(f"[Employee '{self._employee_id}'] Registered {cell.value} in cell {cell.coordinate}")
                break
            
        # Check if event has been correctly registered
        if written:
            # Invalidate evaluated workbook
            self.__invalidate_eval_wb()
        else:
            raise ValueError(f"Event ({event.action} at {event.time}) has not been correctly registered.")

    def commit(self) -> None:
        """
        Commit changes. This must be called after changes have been done (typically after new clock events
        have been registered) to save the modifications.
        """
        # Save the workbook in the local cache
        self._workbook_raw.save(self._file_path)
        LOGGER.debug(f"[Employee '{self._employee_id}'] Saved local file '{self._file_path}'.")

    def evaluate(self) -> None:
        """
        Start an employee's data evaluation. This must be done after changes have been committed.
        Once done, the reading functions are available again and will provide up to date results.
        """
        # After cell values are modified, use libreoffice in headless mode to update the values of formula cells. 
        # This requires 'soffice' to be installed on the system.
        LOGGER.debug(f"[Employee '{self._employee_id}'] Evaluation of '{self._file_path}' started...")
        # Make sure the libre office cache folder exists
        os.makedirs(LIBREOFFICE_CACHE_FOLDER, exist_ok=True)
        # Get temporary file path based on real file path and temporary folder path
        tmp_file = pathlib.Path(LIBREOFFICE_CACHE_FOLDER) / self._file_path.name
        # Ensure the temporary folder doesn't contain a file with the same name
        if tmp_file.exists():
            raise FileExistsError(f"The temporary file '{tmp_file}' already exists and may contain unsaved data. Manual check required.")
        # Create libreoffice command to evaluate and save a spreadsheet
        command = [
            LIBREOFFICE_PATH,                       # Libreoffice invokation
            "--headless",                           # Headless means no GUI
            "--calc",                               # Spreadsheets mode
            "--convert-to", "xlsx",                 # Go to xlsx format
            "--outdir", LIBREOFFICE_CACHE_FOLDER,   # Output in temporary folder
            str(self._file_path.absolute())         # Input employee file
        ]
        # Run command with subprocess
        # Use check=True to ensure raising an exception if the process fails
        # By definition, run() is blocking so it waits for the process to finish
        result = subprocess.run(command, check=True, capture_output=True)
        # It seems like the soffice process can fail and still return 0. Double check for error:
        if result.returncode != 0 or len(result.stderr) > 0:
            raise subprocess.CalledProcessError(returncode=result.returncode, cmd=command, output=result.stderr)
        # Check that the evaluated workbook exists in the temporary folder with the same file name
        if not tmp_file.exists():
            raise FileNotFoundError(f"Subprocess failed to create temporary file under '{tmp_file}'.")
        # Delete old spreadsheet
        self._file_path.unlink(missing_ok=False)
        # Copy temporary spreadsheet to deleted spreadsheet location
        shutil.copy2(src=tmp_file, dst=self._file_path)
        # Delete temporary file
        tmp_file.unlink(missing_ok=True)
        # Load workbook, it will update the readable flag
        self.__load_workbook()
        LOGGER.debug(f"[Employee '{self._employee_id}'] File evaluation succeeded.")

    def close(self) -> None:
        """
        Close the time tracker, save and release resources.
        """
        # Save the employee's file on repository
        self._repository.save_employee_file(self._file_path)
        # Close employee's file
        self._repository.close_employee_file(self._file_path)
        LOGGER.debug(f"[Employee '{self._employee_id}'] Time tracker successfully closed.")

    def __load_workbook(self):
        """
        Load the employee's workbook in raw and evaluated modes. Update the reading flag.
        """
        # Open the employee's workbook in normal (raw) mode to access non evaluated values.
        # Formula cells will contain the raw formula (and not the evaluated value).
        # This workbook can be written and saved safely. It can also be used to read 
        # non formula cell values.
        self._workbook_raw = openpyxl.load_workbook(self._file_path, data_only=False)

        # Open the employee's workbook in data only mode to access evaluated values.
        # If the workbook hasn't been evaluated before, formula cells will contain the
        # value 'None', which will be used to update the reading flag.
        # This workbook is used to read evaluated cell values.
        self._workbook_eval = openpyxl.load_workbook(self._file_path, data_only=True)
        # Check if the workbook is evaluated by checking a formula cell value
        month_sheet = self._workbook_eval.worksheets[self._date.month - 1 + SHEET_JANUARY]
        # Check the first month's day cell arbitrarily
        self._readable = month_sheet[CELL_FIRST_MONTH_DATE].value is not None
        # If not reable, nullify the workbook to prevent unintended access
        if not self._readable:
            self._workbook_eval = None 
        # Log activity
        LOGGER.debug(f"[Employee '{self._employee_id}'] (Re)loaded workbook '{self._file_path}', readable={self._readable}.")

    def __invalidate_eval_wb(self):
        """
        Invalidate the evaluated workbook.
        """
        # Clear readable flag and nullify evaluated workbook
        self._readable = False
        self._workbook_eval = None
