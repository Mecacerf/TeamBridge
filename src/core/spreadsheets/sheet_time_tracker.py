#!/usr/bin/env python3
"""
File: spreadsheet_time_tracker.py
Author: Bastian Cerf
Date: 21/02/2025
Description:
    Implementation of the time tracker analyzer abstract class using
    spreadsheet files for data storage and data analysis. This
    implementation works with the `sheets_repository` module to access
    and manipulate files on a remote repository.

Company: Mecacerf SA
Website: http://mecacerf.ch
Contact: info@mecacerf.ch
"""

# Standard libraries
import logging
import datetime as dt
from typing import TypeVar, Callable, cast

# Third-party libraries
import openpyxl
from openpyxl.utils import column_index_from_string as col_idx
from openpyxl.utils import coordinate_to_tuple
from openpyxl.utils.datetime import from_excel, to_excel  # type: ignore

# Internal imports
from core.time_tracker import *
from .sheets_repository import *
from .libreoffice import *

logger = logging.getLogger(__name__)

# Prevent PIL from spamming debug messages (seems used by openpyxl)
# https://github.com/camptocamp/pytest-odoo/issues/15
logging.getLogger("PIL").setLevel(logging.INFO)

########################################################################
#                   Spreadsheet constants declaration                  #
########################################################################

# Prefix for the evaluated file in the local cache
EVAL_FILE_PREFIX = "eval_"

# Expected spreadsheets major version
# Opening a spreadsheet that doesn't use this major version will fail to
# prevent compatibity issues
# This version may be preceded by a minor version in the form '.xx'
EXPECTED_MAJOR_VERSION = "v250725"

# Init sheet index
SHEET_INIT = 0

# Spreadsheet version
CELL_VERSION = "A3"

# Year the data in the spreasheet belongs to
CELL_YEAR = "A4"

CELL_NAME = "A10"
CELL_FIRSTNAME = "A11"

CELL_OPENING_DAY_SCHEDULE = "A12"
CELL_OPENING_VACATION = "A13"
CELL_OPENING_BALANCE = "A14"
CELL_MAX_CONTINUOUS_WORK_TIME = "A15"

# These are the date and time data analysis is based on
CELL_DATE = "A21"
CELL_TIME = "A22"
CELL_VALIDATION_ANCHOR_DATE = "A23"

# Formula evaluation test cell
CELL_EVALUATED = "A24"

# Error <> description table
CELL_ERROR_TABLE_ID = "A52"  # First error ID row
COL_ERROR_TABLE_DESC = "B"  # Column to get error description
MAX_ERROR_NBR = 10  # Just to give an iteration limit

## Next cells give the locations of different information in the other
## sheets. It allows to support dynamic per sheet locations and improve
## flexibility in production.

# January sheet index
LOC_JANUARY_SHEET = "A25"

# Row number for the first date of the month (01.xx.xxxx)
LOC_FIRST_MONTH_DATE_ROW = "A26"

# Clock in/out times columns (left and right array delimeters)
LOC_FIRST_CLOCK_IN_COL = "A27"
LOC_LAST_CLOCK_OUT_COL = "A28"

# Day related information
LOC_DAY_SCHEDULE_COL = "A29"
LOC_DAY_WORKED_TIME_COL = "A30"
LOC_DAY_BALANCE_COL = "A31"
LOC_DAY_VACATION_COL = "A32"
LOC_DAY_PAID_ABSENCE_COL = "A33"
LOC_DAY_SHEET_ERROR_COL = "A34"
LOC_DAY_SOFT_ERROR_COL = "A35"

# Month related information
LOC_MONTH_SCHEDULE = "A36"
LOC_MONTH_WORKED_TIME = "A37"
LOC_MONTH_BALANCE = "A38"
LOC_MONTH_VACATION = "A39"
LOC_MONTH_PAID_ABSENCE = "A40"

# General information in the month sheets
LOC_EXPECTED_DAY_SCHEDULE = "A41"
LOC_REMAINING_VACATION = "A42"
LOC_YTD_BALANCE = "A43"  # Year-to-date balance
LOC_GLOBAL_ERROR = "A44"

########################################################################
#                           Other constants                            #
########################################################################

T = TypeVar("T")  # For generic methods

CLOSED_ERROR_MSG = "Cannot manipulate a closed tracker."


class _SpecialTime(Enum):
    """Special times the spreadsheet may hold"""

    MIDNIGHT_ROLLOVER = auto()  # Spreadsheet cell value is "24:00"


########################################################################
#               Spreadsheets time tracker implementation               #
########################################################################


class SheetTimeTracker(TimeTrackerAnalyzer):
    """
    Implementation of the time tracker abstract class using spreadsheet
    files for data storage.
    """

    def _setup(self, accessor: SheetsRepoAccessor, readonly: bool):
        """
        Acquire and load the spreadsheet file. Time tracker properties
        are available after this call. The time tracker is never in
        analyzed mode after setup; i.e. the `read_` methods are
        unavalaible.

        `self._workbook_raw` is loaded with `data_only=False`, allowing
        access to raw cell values. In this mode, formula cells contain
        the actual formula text rather than the evaluated result.
        This workbook is always available and can be safely saved.

        If read-only is set:
        - No lock file is created on the remote repository, thus the
        tracker cannot be saved.
        - The raw workbook is opened in read-only, meaning it cannot be
        saved. Therefore the data analysis is not possible and the `read_`
        methods are never available.
        - Only the getters of the `TimeTracker` interface are available.
        - The tracker opening process is about 10 times faster.
        """
        start_ts = time.time()

        self._closed = False
        self._accessor = accessor
        self._readonly = readonly

        # Acquire the spreadsheet file and load the workbook
        self._raw_file_path = accessor.acquire_spreadsheet_file(
            self._employee_id, readonly=readonly
        )
        self._workbook_raw: openpyxl.Workbook = openpyxl.load_workbook(
            self._raw_file_path, data_only=False, read_only=readonly
        )

        # Set the evaluated spreadsheet file path and create the evaluated
        # workbook placeholder
        self._eval_file_path = self._raw_file_path.parent / (
            EVAL_FILE_PREFIX + self._raw_file_path.name
        )
        self._workbook_eval: Optional[openpyxl.Workbook] = None
        self._target_dt = None  # `analyzed` property is False after setup

        # Check that the spreadsheet uses the expected major version
        sheet = self._workbook_raw.worksheets[SHEET_INIT]
        version = sheet[CELL_VERSION].value
        if version is None or not str(version).lower().startswith(
            EXPECTED_MAJOR_VERSION.lower()
        ):
            raise TimeTrackerValueException(
                f"Cannot load workbook '{self._raw_file_path}' that uses version "
                f"'{version}'. The expected major version is "
                f"'{EXPECTED_MAJOR_VERSION}'."
            )

        # Read the locations of the month's sheet cells in the init sheet
        self._sheet_january = int(sheet[LOC_JANUARY_SHEET].value)
        self._row_first_month_date = int(sheet[LOC_FIRST_MONTH_DATE_ROW].value)
        self._col_first_clock_in = col_idx(sheet[LOC_FIRST_CLOCK_IN_COL].value)
        self._col_last_clock_out = col_idx(sheet[LOC_LAST_CLOCK_OUT_COL].value)
        self._col_day_schedule = col_idx(sheet[LOC_DAY_SCHEDULE_COL].value)
        self._col_day_worked_time = col_idx(sheet[LOC_DAY_WORKED_TIME_COL].value)
        self._col_day_balance = col_idx(sheet[LOC_DAY_BALANCE_COL].value)
        self._col_day_vacation = col_idx(sheet[LOC_DAY_VACATION_COL].value)
        self._col_day_paid_absence = col_idx(sheet[LOC_DAY_PAID_ABSENCE_COL].value)
        self._col_day_sheet_error = col_idx(sheet[LOC_DAY_SHEET_ERROR_COL].value)
        self._col_day_soft_error = col_idx(sheet[LOC_DAY_SOFT_ERROR_COL].value)
        self._cell_month_schedule = str(sheet[LOC_MONTH_SCHEDULE].value)
        self._cell_month_worked_time = str(sheet[LOC_MONTH_WORKED_TIME].value)
        self._cell_month_balance = str(sheet[LOC_MONTH_BALANCE].value)
        self._cell_month_vacation = str(sheet[LOC_MONTH_VACATION].value)
        self._cell_month_paid_absence = str(sheet[LOC_MONTH_PAID_ABSENCE].value)
        self._cell_exp_day_schedule = str(sheet[LOC_EXPECTED_DAY_SCHEDULE].value)
        self._cell_remaining_vacation = str(sheet[LOC_REMAINING_VACATION].value)
        self._cell_ytd_balance = str(sheet[LOC_YTD_BALANCE].value)
        self._cell_global_error = str(sheet[LOC_GLOBAL_ERROR].value)

        # Verify that all month sheets are available
        # (spreadsheet integrity check)
        for month in range(13):
            self._workbook_raw.worksheets[self.__get_month_sheet_idx(month)]

        delta_ts = (time.time() - start_ts) * 1000.0

        logger.debug(
            f"[Employee '{self._employee_id}'] Spreadsheet time "
            f"tracker setup in {delta_ts:.0f}ms "
            f"(read-only={self._readonly})."
        )

    ## Evaluated workbook loading and closing methods ##

    def __close_eval_workbook(self):
        """
        Close the evaluated workbook, if opened.

        This call doesn't reset the `target_dt` property; i.e. the
        `analyzed` flag is not cleared after this call. Prefer using
        `self.__invalidate_analysis()` unless really intended.
        """
        if self._workbook_eval is None:
            return  #  Already closed

        try:
            self._workbook_eval.close()
            self._eval_file_path.unlink()
        except Exception:
            # This is unlikely to happen and probably not critical in most
            # cases, just log the exception and try to continue.
            logger.error(
                f"Error occurred closing evaluated workbook '{self._eval_file_path}'.",
                exc_info=True,
            )
        finally:
            self._workbook_eval = None

    def __invalidate_analysis(self):
        """
        Invalidate the analysis. It closes the evaluated workbook and
        clear the `target_dt` property (=`None`).
        After this call, the `analyzed` property is `False` and the
        `read_` method are unavailable.
        """
        self.__close_eval_workbook()
        self._target_dt = None

    def __save_workbook(self):
        """
        Save the raw workbook in the local cache.

        Raises:
            TimeTrackerSaveException: General saving error.
        """
        assert not self._closed, CLOSED_ERROR_MSG
        assert not self._readonly  # This check must raise earlier

        try:
            self._workbook_raw.save(self._raw_file_path)
        except Exception as e:
            raise TimeTrackerSaveException() from e

        # Automatically invalidate analysis after save: raw data have
        # changed and a new analysis is required to read up-to-date values.
        self.__invalidate_analysis()

    ## Utility methods to find sheet or row indexes ##

    def __get_date_row(self, date: dt.date | dt.datetime) -> int:
        """
        Args:
            date (date | datetime): Input date.

        Returns:
            int: The row index for the given date in any month sheet.

        Raises:
            TimeTrackerDateException: Date is outside `tracked_year`.
        """
        if date.year != self.tracked_year:
            raise TimeTrackerDateException(
                f"Date '{date}' is outside of tracked year {self.tracked_year}."
            )

        return date.day - 1 + self._row_first_month_date

    def __get_month_sheet_idx(self, month: int | dt.date | dt.datetime) -> int:
        """
        Args:
            month (int | date | datetime): Input month.

        Returns:
            int: The sheet index for to the given month.

        Raises:
            TimeTrackerDateException: Date is outside `tracked_year`.
        """
        if isinstance(month, (dt.date, dt.datetime)):
            if month.year != self.tracked_year:
                raise TimeTrackerDateException(
                    f"Date '{month}' is outside of tracked year {self.tracked_year}."
                )
            month = month.month

        return month - 1 + self._sheet_january

    def __get_clock_action_for_cell(self, cell: Any) -> ClockAction:
        """
        Get the clock action (clock-in / clock-out) that the cell is
        supposed to hold. The function assumes the sequence
        `self._col_first_clock_in` is a clock-in;
        `self._col_first_clock_in + 1` is a clock-out;
        `self._col_first_clock_in + 2` is a clock-in; etc until
        `self._col_last_clock_out`

        Args:
            cell (Any): Openpyxl cell.

        Returns:
            ClockAction: Clock action for the cell.
        """
        assert self._col_first_clock_in <= cell.column <= self._col_last_clock_out

        # Follow the sequence 0: clock-in, 1: cock-out, 2: clock-in, etc.
        even = ((cell.column - self._col_first_clock_in) % 2) == 0
        return ClockAction.CLOCK_IN if even else ClockAction.CLOCK_OUT

    ## Utility methods to get cell values ##

    def __get_init_cell_val(self, cell: str, cast_func: Callable[[Any], T]) -> T:
        """
        Get a cell value from the init sheet with type checking.

        Args:
            cell (str): Cell coordinate (such as 'A8').
            cast_func (Callable[[Any], T]): Type casting function.

        Returns:
            T: Value of expected type.

        Raises:
            TimeTrackerValueException: Conversion unsupported for the
                given value.
        """
        assert not self._closed, CLOSED_ERROR_MSG

        value = self._workbook_raw.worksheets[SHEET_INIT][cell].value
        return self.__cast(value, cast_func)

    def __get_month_cell_value(
        self,
        workbook: openpyxl.Workbook,
        month: int | dt.date | dt.datetime,
        cell: str,
        cast_func: Callable[[Any], T],
    ) -> T:
        """
        Get a cell value from a month sheet with type checking.

        Args:
            workbook (openpyxl.Workbook): The workbook to read.
            month (int | date | datetime): Month information.
            cell (str): Cell coordinate (such as 'A8').
            cast_func (Callable[[Any], T]): Type casting function.

        Returns:
            T: Value of expected type.

        Raises:
            TimeTrackerValueException: Conversion unsupported for the
                given value.
            TimeTrackerDateException: Date is outside `tracked_year`.
        """
        assert not self._closed, CLOSED_ERROR_MSG

        sheet = workbook.worksheets[self.__get_month_sheet_idx(month)]
        return self.__cast(sheet[cell].value, cast_func)

    def __get_month_day_cell_value(
        self,
        workbook: openpyxl.Workbook,
        date: dt.date | dt.datetime,
        col: int,
        cast_func: Callable[[Any], T],
    ) -> T:
        """
        Get a cell value from a month sheet with type checking. The cell
        row is calculated based on the given date.

        Args:
            workbook (openpyxl.Workbook): The workbook to read.
            date (date | datetime): Day and month information.
            col (int): Data column.
            cast_func (Callable[[Any], T]): Type casting function.

        Returns:
            T: Value of expected type.

        Raises:
            TimeTrackerValueException: Conversion unsupported for the
                given value.
            TimeTrackerDateException: Date is outside `tracked_year`.
        """
        assert not self._closed, CLOSED_ERROR_MSG

        day_row = self.__get_date_row(date)
        sheet = workbook.worksheets[self.__get_month_sheet_idx(date)]
        value = sheet.cell(row=day_row, column=col).value
        return self.__cast(value, cast_func)

    def __read_month_cell_value(
        self,
        month: int | dt.date | dt.datetime | None,
        cell: str,
        cast_func: Callable[[Any], T],
    ) -> T:
        """
        Same as `self.__get_month_cell_value()` but the value is read in
        the evaluated workbook, if available.

        Raises:
            TimeTrackerReadException: Time tracker not analyzed.
            TimeTrackerValueException: Conversion unsupported for the
                given value.
            TimeTrackerDateException: Date is outside `tracked_year`.
        """
        # Allowing `None` month allows to directly pass `target_dt` as argument
        # without complaint from the type checker
        if self._workbook_eval is None or month is None:
            raise TimeTrackerReadException("The time tracker must be analyzed.")

        return self.__get_month_cell_value(self._workbook_eval, month, cell, cast_func)

    def __read_month_day_cell_value(
        self,
        date: dt.date | dt.datetime | None,
        col: int,
        cast_func: Callable[[Any], T],
    ) -> T:
        """
        Same as `self.__get_month_day_cell_value()` but the value is read
        in the evaluated workbook, if available.

        Raises:
            TimeTrackerReadException: Time tracker not analyzed.
            TimeTrackerValueException: Conversion unsupported for the
                given value.
            TimeTrackerDateException: Date is outside `tracked_year`.
        """
        # Allowing `None` date allows to directly pass `target_dt` as argument
        # without complaint from the type checker
        if self._workbook_eval is None or date is None:
            raise TimeTrackerReadException("The time tracker must be analyzed.")

        return self.__get_month_day_cell_value(
            self._workbook_eval, date, col, cast_func
        )

    ## Utility methods to set cell values ##

    def __set_month_day_cell_value(
        self, date: dt.date | dt.datetime, col: int, value: Any
    ):
        """
        Set a cell value in the month sheet. The cell row is calculated
        based on the given date.

        Args:
            date (date | datetime): Day and month information.
            col (int): Data column.
            value (Any): Value to set.

        Raises:
            TimeTrackerDateException: Date is outside `tracked_year`.
            TimeTrackerWriteException: Time tracker is closed.
            TimeTrackerSaveException: General saving error.
        """
        assert not self._closed, CLOSED_ERROR_MSG

        if self._readonly:
            raise TimeTrackerWriteException("Cannot write in a read-only tracker.")

        date_row = self.__get_date_row(date)
        sheet_idx = self.__get_month_sheet_idx(date)
        month_sheet = self._workbook_raw.worksheets[sheet_idx]

        cell = month_sheet.cell(row=date_row, column=col)
        cell.value = value

        self.__save_workbook()

    ## Utility method for type safety ##

    def __cast(self, value: Any, cast_func: Callable[[Any], T]) -> T:
        """
        Cast / convert the given value to the expected type.

        Args:
            value (Any): The value to cast / convert.
            cast_func (Callable[[Any], T]): Casting function.

        Returns:
            T: The casted / converted value.

        Raises:
            TimeTrackerValueException: Conversion unsupported for the
                given value.
        """
        try:
            value = cast_func(value)
        except (TypeError, ValueError) as e:
            raise TimeTrackerValueException() from e

        return value

    def __to_timedelta(self, value: Any) -> dt.timedelta:
        """
        Convert the given value to a `datetime.timedelta` if necessary.

        Args:
            value (Any): Input value.

        Returns:
            datetime.timedelta: Converted value.
        """
        if isinstance(value, (float, int)):
            # Handle specific case where the value has been set but not reparsed
            # by openpyxl, which results in a time still being represented as
            # a fraction of days (a number).
            value = cast(Any, from_excel(value))

        if isinstance(value, dt.timedelta):
            # Passthrough
            return value

        if isinstance(value, (dt.time, dt.datetime)):
            return dt.timedelta(
                hours=value.hour, minutes=value.minute, seconds=value.second
            )

        raise ValueError(f"Cannot convert {type(value).__name__} to timedelta.")

    def __to_time(self, value: Any) -> dt.time | _SpecialTime:
        """
        Convert the given value to a `datetime.time` if necessary.

        Args:
            value (Any): Input value.

        Returns:
            Union[dt.time, _SpecialTime]: Converted value as a `dt.time`
                or a `_SpecialTime` value.
        """
        if isinstance(value, (float, int)):
            # Handle specific case where the value has been set but not reparsed
            # by openpyxl, which results in a time still being represented as
            # a fraction of days (a number).
            value = cast(Any, from_excel(value))

        if isinstance(value, dt.timedelta):
            # A timedelta can be converted to a time or datetime by converting
            # it to excel fromat (fraction of days) and back again to time.
            value = cast(Any, from_excel(to_excel(value), timedelta=False))

        if isinstance(value, dt.time):
            # Passthrough
            return value

        if isinstance(value, dt.datetime):
            # The midnight rollover, noted 24:00 in the cell, is interpreted by
            # Openpyxl as midnight on the 01.01.1900. While it makes sense that
            # the time is midnight, it's not so clear why the date is
            # 01.01.1900. To prevent compatibility issues, only check that the
            # datetime refers to midnight and assume it is a midnight rollover
            # value.
            if value.hour == 0 and value.minute == 0:
                return _SpecialTime.MIDNIGHT_ROLLOVER

            # Other datetimes are not supported.

        raise ValueError(f"Cannot convert {type(value).__name__} to time.")

    def __to_date(self, value: Any) -> dt.date:
        """
        Convert the given value to a date, if necessary.

        Args:
            value (Any): Input value.

        Returns:
            dt.date: Converted value.
        """
        if isinstance(value, dt.datetime):
            return value.date()

        if isinstance(value, dt.date):
            return value

        raise ValueError(f"Cannot convert {type(value).__name__} to date.")

    def __to_float_none_safe(self, value: Any) -> float:
        """
        Convert the given value to a float. A `None` value is interpreted
        as 0.0.

        Args:
            value (Any): Input value.

        Returns:
            float: Converted value.
        """
        if value is None:
            return 0.0

        if isinstance(value, (float, int)):
            return float(value)

        raise ValueError(f"Cannot convert {type(value).__name__} to float.")

    def __to_int_none_safe(self, value: Any) -> int:
        """
        Convert the given value to an integer. A `None` value is
        interpreted as 0.

        Args:
            value (Any): Input value.

        Returns:
            int: Converted value.
        """
        if value is None:
            return 0

        if isinstance(value, int):
            return int(value)

        raise ValueError(f"Cannot convert {type(value).__name__} to int.")

    def __to_str_none_safe(self, value: Any) -> str:
        """
        Convert the given value to a string. A `None` value is
        interpreted as an empty string.

        Args:
            value (Any): Input value.

        Returns:
            str: Converted value.
        """
        if value is None:
            return ""

        if isinstance(value, str):
            return value

        raise ValueError(f"Cannot convert {type(value).__name__} to str.")

    ## Utility methods ##

    def __get_clock_event(self, cell: Any) -> Optional[ClockEvent]:
        """
        Try to convert the cell content to a `ClockEvent`. The cell must
        contain a time in the `hh:mm` format or must be empty.

        Args:
            cell (Any): Openpyxl cell.

        Returns:
            Optional[ClockEvent]: A `ClockEvent` if the cell has a parsable
                value or `None` otherwise.

        Raises:
            TimeTrackerValueException: Conversion unsupported for the
                cell value.
        """
        if cell is None or cell.value is None:
            return None  # Cell is empty, no clock event registered

        evt_time = self.__cast(cell.value, self.__to_time)

        if isinstance(evt_time, dt.time):
            # Standard clock-in / clock-out
            action = self.__get_clock_action_for_cell(cell)
            clk_evt = ClockEvent(evt_time, action)
        else:
            # Handle a special time value
            if evt_time == _SpecialTime.MIDNIGHT_ROLLOVER:
                clk_evt = ClockEvent.midnight_rollover()
            else:
                raise TimeTrackerValueException(f"Unhandled special time {evt_time}.")

        return clk_evt

    def __set_clock_event(self, cell: Any, event: ClockEvent):
        """
        Set the cell value based on the given `ClockEvent`.

        Changes are not saved.

        Args:
            cell (Any): Openpyxl cell to write into.
            event (ClockEvent): ClockEvent to write in the cell.
        """
        if event is ClockEvent.midnight_rollover():
            # Write the special time value "24:00"
            cell.value = 1.0  # One full day
        else:
            # Write a standard time value
            cell.value = dt.time(
                hour=event.time.hour, minute=event.time.minute, second=0
            )

    ## General employee's properties access ##

    @property
    def firstname(self) -> str:
        return self.__get_init_cell_val(CELL_FIRSTNAME, str)

    @property
    def name(self) -> str:
        return self.__get_init_cell_val(CELL_NAME, str)

    ## General time tracker's properties access ##

    @property
    def tracked_year(self) -> int:
        return self.__get_init_cell_val(CELL_YEAR, int)

    @property
    def opening_day_schedule(self) -> dt.timedelta:
        return self.__get_init_cell_val(CELL_OPENING_DAY_SCHEDULE, self.__to_timedelta)

    @property
    def opening_balance(self) -> dt.timedelta:
        return self.__get_init_cell_val(CELL_OPENING_BALANCE, self.__to_timedelta)

    @property
    def opening_vacation_days(self) -> float:
        return self.__get_init_cell_val(CELL_OPENING_VACATION, float)

    @property
    def max_clock_events_per_day(self) -> int:
        return int(self._col_last_clock_out - self._col_first_clock_in + 1)

    @property
    def max_continuous_work_time(self) -> dt.timedelta:
        return self.__get_init_cell_val(
            CELL_MAX_CONTINUOUS_WORK_TIME, self.__to_timedelta
        )

    ## Time Tracker read / write methods ##

    def get_clocks(self, date: dt.date | dt.datetime) -> list[Optional[ClockEvent]]:
        """
        Implementation of `TimeTracker.get_clocks()`.

        Retrieve the clock events for a specific date by iterating through
        the corresponding row in the monthly sheet for that date.

        Raises:
            TimeTrackerDateException: Date is outside `tracked_year`.
            TimeTrackerValueException: Read an unexpected value in the row.
        """
        assert not self._closed, CLOSED_ERROR_MSG

        date_row = self.__get_date_row(date)
        sheet_idx = self.__get_month_sheet_idx(date)
        month_sheet = self._workbook_raw.worksheets[sheet_idx]

        clock_events: list[Optional[ClockEvent]] = []

        # Iterate columns along the date row from the first clock-in column to
        # the last clock-out column.
        # This is to avoid using iter_cols() which isn't supported in read-only
        # mode.
        row = next(
            month_sheet.iter_rows(
                min_row=date_row,
                max_row=date_row,
                min_col=self._col_first_clock_in,
                max_col=self._col_last_clock_out,
            )
        )

        # Transpose single row to get columns as single-cell tuples
        columns = zip(row)
        for column in columns:
            # The column contains a single row (min_row = max_row)
            cell = column[0]
            event: Optional[ClockEvent] = None

            # Create and append a ClockEvent if a time is available in the cell,
            # otherwise just append None in the clock events list
            event = self.__get_clock_event(cell)

            clock_events.append(event)

        # Return the events list without the trailing None values
        for i in reversed(range(len(clock_events))):
            if clock_events[i] is not None:
                return clock_events[: i + 1]
        return []  # All values are None

    def register_clock(self, date: dt.date | dt.datetime, event: ClockEvent):
        """
        Implementation of `TimeTracker.register_clock()`.

        Searches for the next available time slot in the current date's
        row on the month sheet and writes the event time. If no slot is
        available, raises a TimeTrackerWriteException.

        Raises:
            TimeTrackerDateException: Date is outside `tracked_year`.
            TimeTrackerWriteException: No available slot found to write
                the clock event.
            TimeTrackerSaveException: Unable to save the tracker in the
                local cache.
        """
        assert not self._closed, CLOSED_ERROR_MSG

        if self._readonly:
            raise TimeTrackerWriteException(
                "Cannot register event in read-only tracker."
            )

        date_row = self.__get_date_row(date)
        sheet_idx = self.__get_month_sheet_idx(date)
        month_sheet = self._workbook_raw.worksheets[sheet_idx]

        # Determine starting column index based on clock action
        start_idx = self._col_first_clock_in + (
            1 if event.action == ClockAction.CLOCK_OUT else 0
        )

        for col_idx in range(start_idx, self._col_last_clock_out + 1, 2):
            cell = month_sheet.cell(row=date_row, column=col_idx)

            if cell.value is None:
                # The slot is available, write the clock time in HH:MM format
                self.__set_clock_event(cell, event)

                # Save the workbook and automatically invalidate a previous
                # data analysis
                self.__save_workbook()

                logger.debug(
                    f"[Employee '{self._employee_id}'] Registered clock event "
                    f"in sheet '{sheet_idx}' : "
                    f"'{event}' in '{cell.coordinate}'."
                )

                return

        raise TimeTrackerWriteException(
            f"No available time slot to write the clock event '{event}'."
        )

    def write_clocks(
        self, date: dt.date | dt.datetime, events: list[Optional[ClockEvent]]
    ):
        """
        Implementation of `TimeTracker.write_clocks()`.

        Raises:
            TimeTrackerDateException: Date is outside `tracked_year`.
            TimeTrackerWriteException: Too many events.
            TimeTrackerSaveException: Unable to save the tracker in the
                local cache.
        """
        assert not self._closed, CLOSED_ERROR_MSG

        if self._readonly:
            raise TimeTrackerWriteException("Cannot write events in read-only tracker.")

        date_row = self.__get_date_row(date)
        sheet_idx = self.__get_month_sheet_idx(date)
        month_sheet = self._workbook_raw.worksheets[sheet_idx]

        if len(events) > self.max_clock_events_per_day:
            raise TimeTrackerWriteException(
                f"List exceeds {self.max_clock_events_per_day} events."
            )

        msgs: list[str] = []

        # Whatever the events list size, all cells are overridden.
        # Note: the range stop value is exclusive while last clock out column
        # index is inclusive, reason why one is added.
        evts_rng = range(self._col_first_clock_in, self._col_last_clock_out + 1)
        # Pads with None events
        clk_evts = events + [None] * (len(evts_rng) - len(events))

        assert len(evts_rng) == self.max_clock_events_per_day

        for col_idx, evt in zip(evts_rng, clk_evts):
            cell = month_sheet.cell(row=date_row, column=col_idx)

            if evt:
                self.__set_clock_event(cell, evt)
            else:
                cell.value = None

            msgs.append(f"'{evt}' in '{cell.coordinate}'")

        # Save the workbook and automatically invalidate a previous
        # data analysis
        self.__save_workbook()

        logger.debug(
            f"[Employee '{self._employee_id}'] Written clock events "
            f"in sheet '{sheet_idx}' : "
            f"{" | ".join(msgs)}"
        )

    def set_vacation(self, date: dt.date | dt.datetime, day_ratio: float):
        self.__set_month_day_cell_value(date, self._col_day_vacation, day_ratio)

    def get_vacation(self, date: dt.date | dt.datetime) -> float:
        return self.__get_month_day_cell_value(
            self._workbook_raw, date, self._col_day_vacation, self.__to_float_none_safe
        )

    def set_paid_absence(self, date: dt.date | dt.datetime, day_ratio: float):
        self.__set_month_day_cell_value(date, self._col_day_paid_absence, day_ratio)

    def get_paid_absence(self, date: dt.date | dt.datetime) -> float:
        return self.__get_month_day_cell_value(
            self._workbook_raw,
            date,
            self._col_day_paid_absence,
            self.__to_float_none_safe,
        )

    def set_attendance_error(self, date: dt.date | dt.datetime, error_id: int):
        self.__set_month_day_cell_value(
            date, self._col_day_soft_error, error_id or None
        )

    def get_attendance_error(self, date: dt.date | dt.datetime) -> int:
        return self.__get_month_day_cell_value(
            self._workbook_raw, date, self._col_day_soft_error, self.__to_int_none_safe
        )

    def get_attendance_error_desc(self, error_id: int) -> str:
        """
        Look up the description corresponding to a given error ID
        in the error ID <> description table of the raw workbook.
        """
        assert not self._closed, CLOSED_ERROR_MSG

        start_row, col_id = coordinate_to_tuple(CELL_ERROR_TABLE_ID)
        col_desc = col_idx(COL_ERROR_TABLE_DESC)

        sheet = self._workbook_raw.worksheets[SHEET_INIT]

        for row in sheet.iter_rows(
            min_row=start_row,
            max_row=start_row + MAX_ERROR_NBR,
            min_col=col_id,
            max_col=col_desc,
            values_only=True,
        ):
            current_id = row[0]

            if current_id is None:
                break

            current_id = self.__cast(current_id, int)
            if current_id == error_id:
                description = row[col_desc - col_id]
                return self.__cast(description, self.__to_str_none_safe)

        raise TimeTrackerValueException(
            f"No description found for error ID {error_id}."
        )

    def set_last_validation_anchor(self, date: dt.date):
        """
        Write the date in the init sheet cell. The workbook is saved
        without invalidating the analysis results because this information
        is not related to the analysis data.
        """
        if self.tracked_year != date.year:
            raise TimeTrackerDateException()

        init_sheet = self._workbook_raw.worksheets[SHEET_INIT]
        init_sheet[CELL_VALIDATION_ANCHOR_DATE].value = self.__cast(
            date, self.__to_date
        )

        try:
            self._workbook_raw.save(self._raw_file_path)
        except Exception as e:
            raise TimeTrackerSaveException() from e

    def get_last_validation_anchor(self) -> dt.date:
        return self.__get_init_cell_val(CELL_VALIDATION_ANCHOR_DATE, self.__to_date)

    ## Time Tracker Analyzer methods ##

    def _analyze_internal(self):
        """
        Implementation of `TimeTrackerAnalyzer.analyze()`.

        Evaluate the spreadsheet file at the specified target date and
        time using the LibreOffice module.

        The evaluation is performed on a copy of the main spreadsheet file
        to avoid modifying the default "date" and "time" fields in the
        init sheet. This also allows the file to be loaded in `read_only`
        mode, which improves performance but requires keeping a process
        open on the file. The file is loaded with `data_only=True`,
        enabling access to evaluated formula values.

        After loading the workbook, a check is performed to determine
        whether the formula cells have actually been evaluated. If not,
        a `TimeTrackerAnalysisException` is raised.

        Raises:
            RuntimeError: No LibreOffice installation found.
            FileExistsError: A previous evaluation didn't finish properly
                and a file is still existing in the temporary folder.
            RuntimeError: The LibreOffice execution returned an error.
            TimeoutError: The LibreOffice execution timed out.
            FileNotFoundError: LibreOffice didn't produce the expected output.
            TimeTrackerAnalysisException: The produced file wasn't evaluated.

        Exceptions are chained to the more general `TimeTrackerAnalysisException`
        by the `analyze()` method.
        """
        assert not self._closed, CLOSED_ERROR_MSG
        assert self._target_dt is not None  # Set by analyze()

        if self._readonly:
            raise TimeTrackerAnalysisException("Cannot analyze a read-only tracker")

        start_ts = time.time()

        # Step 1: close old evaluated workbook file and delete previous
        # evaluated file path (if existing)
        self.__close_eval_workbook()

        # Step 2: Save a copy of the raw workbook with date and time set to
        # target datetime.
        init_sheet = self._workbook_raw.worksheets[SHEET_INIT]
        tmp_date_formula = init_sheet[CELL_DATE].value
        tmp_time_formula = init_sheet[CELL_TIME].value

        try:
            init_sheet[CELL_DATE].value = self._target_dt.date()
            init_sheet[CELL_TIME].value = self._target_dt.time()
            self._workbook_raw.save(self._eval_file_path)

        finally:
            # Always restore the formulas in the init sheet
            init_sheet[CELL_DATE].value = tmp_date_formula
            init_sheet[CELL_TIME].value = tmp_time_formula

        # Step 3: evaluate the workbook copy using the libreoffice module
        try:
            # On success, this call replaces the given file with an evaluated
            # one
            evaluate_calc(self._eval_file_path)

            self._workbook_eval = openpyxl.load_workbook(
                self._eval_file_path, data_only=True, read_only=True
            )

            # Read the "evaluated" cell from the init sheet to check whether
            # the formula values are available
            if self._workbook_eval.worksheets[SHEET_INIT][CELL_EVALUATED].value is None:
                raise TimeTrackerAnalysisException()

        except Exception:
            # Reset the state to step 1
            # target_dt is set to None by the superclass on exception, which
            # sets the `analyzed` property to False
            self.__close_eval_workbook()
            raise

        finally:
            delta_ts = (time.time() - start_ts) * 1000.0
            logger.debug(
                f"[Employee '{self._employee_id}'] "
                + (
                    f"Data analysis failed after {delta_ts:.0f}ms."
                    if self._workbook_eval is None
                    else f"Data analyzed in {delta_ts:.0f}ms."
                )
            )

    def read_day_schedule(self, date: dt.date | dt.datetime) -> dt.timedelta:
        return self.__read_month_day_cell_value(
            date, self._col_day_schedule, self.__to_timedelta
        )

    def read_day_worked_time(self, date: dt.date | dt.datetime) -> dt.timedelta:
        return self.__read_month_day_cell_value(
            date, self._col_day_worked_time, self.__to_timedelta
        )

    def read_day_balance(self, date: dt.date | dt.datetime) -> dt.timedelta:
        return self.__read_month_day_cell_value(
            date, self._col_day_balance, self.__to_timedelta
        )

    def read_day_attendance_error(self, date: dt.date | dt.datetime) -> int:
        return self.__read_month_day_cell_value(
            date, self._col_day_sheet_error, self.__to_int_none_safe
        )

    def read_month_expected_daily_schedule(
        self, month: int | dt.date | dt.datetime
    ) -> dt.timedelta:
        return self.__read_month_cell_value(
            month, self._cell_exp_day_schedule, self.__to_timedelta
        )

    def read_month_schedule(self, month: int | dt.date | dt.datetime) -> dt.timedelta:
        return self.__read_month_cell_value(
            month, self._cell_month_schedule, self.__to_timedelta
        )

    def read_month_worked_time(
        self, month: int | dt.date | dt.datetime
    ) -> dt.timedelta:
        return self.__read_month_cell_value(
            month, self._cell_month_worked_time, self.__to_timedelta
        )

    def read_month_balance(self, month: int | dt.date | dt.datetime) -> dt.timedelta:
        return self.__read_month_cell_value(
            month, self._cell_month_balance, self.__to_timedelta
        )

    def read_month_vacation(self, month: int | dt.date | dt.datetime) -> float:
        return self.__read_month_cell_value(month, self._cell_month_vacation, float)

    def read_year_vacation(self) -> float:
        return self.opening_vacation_days - self.read_year_remaining_vacation()

    def read_year_remaining_vacation(self) -> float:
        # Warning: to get the remaining vacation days for the whole year, read
        # the value from the December sheet. This way the value is updated
        # even if vacations are planned for future months.
        return self.__read_month_cell_value(12, self._cell_remaining_vacation, float)

    def read_year_to_date_balance(self) -> dt.timedelta:
        return self.__read_month_cell_value(
            self._target_dt, self._cell_ytd_balance, self.__to_timedelta
        )

    def read_year_attendance_error(self) -> int:
        # As for the vacations, read the December sheet that contains the very
        # last error value.
        return self.__read_month_cell_value(
            12, self._cell_global_error, self.__to_int_none_safe
        )

    def save(self):
        """
        Implementation of `TimeTracker.save()`.

        Try to save the file on the spreadsheets repository.

        Raises:
            TimeTrackerSaveException: General saving error.
            See chained exception for details.
        """
        assert not self._closed, CLOSED_ERROR_MSG

        if self._readonly:
            raise TimeTrackerSaveException("Cannot save a read-only tracker.")

        try:
            # Only save if in read/write mode
            self._accessor.save_spreadsheet_file(self._raw_file_path)

        except Exception as e:
            raise TimeTrackerSaveException() from e

        logger.debug(f"[Employee '{self._employee_id}'] Spreadsheet file saved.")

    def close(self):
        """
        Implementation of `TimeTracker.close()`.

        Release the file lock from the spreadsheets repository, allowing
        other processes to access the employee's file. Files are also
        deleted from the local cache.

        Raises:
            TimeTrackerCloseException: General error while closing.
            See chained exception for details.
        """
        assert not self._closed, CLOSED_ERROR_MSG

        self.__invalidate_analysis()  # Close the evaluated workbook

        try:
            self._workbook_raw.close()  # Required if opened in read-only mode
            self._accessor.release_spreadsheet_file(
                self._raw_file_path, readonly=self._readonly
            )

        except Exception as e:
            raise TimeTrackerCloseException() from e
        finally:
            # Even if an error occurs, prevent any further access because the
            # file is in an unknown state.
            self._closed = True

        logger.debug(
            f"[Employee '{self._employee_id}'] Spreadsheet time tracker closed."
        )
